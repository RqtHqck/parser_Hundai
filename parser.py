# BASE LIBRARY
import random
import threading
import time
import os
import json
import re
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import SheetTitleException
# THREADING
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from multiprocessing import Pool
import multiprocessing
# REQUEST&BS4
import requests
from bs4 import BeautifulSoup
import fake_useragent
# SELENIUM
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
# MY IMPORTS
from settings.ParserClass import Parser

URL = 'https://www.elcats.ru'   #URL сайта
MODELS_URL = "https://www.elcats.ru/hyundai/"   #URL с модельным рядом
AUTO_BRAND = 'HUNDAI'   #Бренд
AUTO_MODEL = None   #Выбранная модель
MODEL_URL = None    #Ссылка на выбранную модель
parser = Parser()
class Time(Parser):
    def __init__(self):
        self.time_start = time.time()  # Сохранение начального времени
    
    def end(self):
        time_end = time.time()
        elapsed_time = time_end - self.time_start  # Вычисление затраченного времени
        self.logger(f'|--- Сбор всех данных занял {elapsed_time:.2f} секунд.')  # Логирование времени
            
        
def parse_all_models_into_file(url, parser):
    """
    Парсит все модели со страницы моделей HUNDAI
    Возвращает список models и список urls,
    """
    try:
        # Находим таблицу
        soup = BeautifulSoup(parser.fetch_data(url).text, 'lxml') 
        tables = soup.find_all('table', class_="parts_table")
        # Получаем модели и юрл к ним
        models = [link.text.strip() for table in tables for link in table.find_all('a') if link.text.strip()]
        urls = [f'https://www.elcats.ru/hyundai/Modification.aspx?Model={model}' for model in models]
        
        # Сохраняем найденные ресурсы в JSON файл
        source = {}
        for model, url in zip(models, urls):
            source[model] = url
        parser.save_data(name="Models", path='data', src=source)
        
        return [models, urls]
    except Exception as e:
        parser.logger(f'Ошибка при выполнении функции parse_all_models_into_file: {e}')
        raise
    

def choose_model_by_user(args):
    """
    Выбор модели бренда пользователем.
    Возвращает модель и ссылку на модель
    """
    try:
        # Отрисовываем варианты моделей для выбора из USER_CHOOSE.env
        models, urls = args
        for i, (model, url) in enumerate(zip(models, urls), start=1):
            print(f"[#{i}]: {model}")
    except Exception as e:
        print(f'Ошибка при выполнении функции choose_model_by_user: {e}')
        raise
    
    # Для сервера используем предвыбор
    load_dotenv(dotenv_path='settings/USER_CHOOSE.env')
    choice = int(os.getenv('KEY', 1)) -1

    try:
        model = models[choice]
        model_url = urls[choice]
        return model, model_url
    except Exception as e:
        print(f'Ошибка при выполнении функции choose_model_by_user: {e}')
        raise

    # Для пользователя используем выбор
    # while True:
    #     try:
    #         # Предлагаем выбор   
    #         choice = int(input('Выберите номер модели из предложенных: ')) - 1
    #         # Проверка корректности выбора модели
    #         if 0 <= choice < len(models):
    #             model = models[choice]
    #             model_url = urls[choice]
    #             print(f"\nВыбранная модель: [#{choice + 1}]: {model}")
    #             return model, model_url
    #         else:
    #             logger("Некорректный номер. Попробуйте снова.", saveonly=False, first=False, infunction=True)
    #     except (ValueError, IndexError):
    #         logger("Некорректный номер. Введите числовое значение.", saveonly=False, first=False, infunction=True)
    #     except Exception as e:
    #         logger(f'Ошибка при выполнении функции choose_model_by_user: {e}', saveonly=False, first=False, infunction=True)
    #         raise
 
                
def parse_region(url, parser):
    """
    Парсит всё в регионе (модификации)
    Возвращает 
    """
    try:
        soup = BeautifulSoup(parser.fetch_data(url).text, 'lxml')
        table = soup.find('div', id="content").find('fieldset').find('table', cellpadding="6").find_all('tr')
        regions = table[1].find_all('td')
        table = table[3:]

        # Проверка есть ли регион европа
        regions = [item.find('label').text for item in regions]
        if regions and 'ЕВРОПА' in regions:
            js_functions = [item.find('a')['href'] for item in table]
            years = [year.find_all('td')[-1].text.strip() for year in table if year]
            regex = re.compile(r"submit\('([^']*)','([^']*)'\)")
            url_list = []
            models_name = []
            for js in js_functions:
                match = regex.search(js)
                if match:
                    arg1 = match.group(1)
                    arg2 = match.group(2)
                    region_url = f"https://www.elcats.ru/hyundai/Options.aspx?Code={arg1}&Title={arg2}"
                    url_list.append(region_url)
                    models_name.append(arg2)
            return models_name, url_list, years
        raise LookupError('В выбранной модели нету региона "ЕВРОПА"')
    except Exception as e:
        parser.logger(f'Ошибка при выполнении функции parse_region: {e}', saveonly=False, first=False, infunction=True)
        raise


def detail_process(task, lock, parser):
    parser.logger('---')
    # Настройка драйвера и парсера
    driver = parser.setup_driver()
    # Получает аргументы из потока
    model_name, model_url, issue_date = task
    # Ищет url комплектации
    details_url = parse_complectation(model_url, driver, parser)  # Комплектация
    details_dict = parse_details_page(details_url, parser)

    regex_for_cut_sub_title =  re.compile(r"submit\('([^']*)','([^']*)','([^']*)','([^']*)'\)")
    for group, details_js in details_dict.items():
        # Перебор деталей для перехода на страницу part_url с таблицей
        for detail_js in details_js:
            try:
                # Part_url - переход
                conditional_part_url = parser.selenium_crossing(details_url, detail_js, driver)
                # Очистка cut_sub_title
                match = regex_for_cut_sub_title.search(detail_js)
                cut_sub_title = match.group(3) if match else '-'
                # print(conditional_part_url)
            except:
                parser.logger(f'Ошибка при переходе по js детали на страницу с таблицей запчасти')

            # part_mage_url

            # Добавяем соуществующие данные в year_data
            year_data = [AUTO_BRAND, model_name, issue_date, cut_sub_title, model_url]
            if 'Parts.aspx' in conditional_part_url:
                # Если страница с таблицей парсим данные (внутри вызывает 2 функции)
                part_url = conditional_part_url
                if part_url is None:
                    continue
                part_mage_url = parse_part_picture(URL, conditional_part_url, parser)
                year_data.append(part_url)
                year_data.append(part_mage_url)
                cleared_model_name = clear_model_name_for_book(model_name, parser)

                parse_table(part_url, year_data, cleared_model_name, lock, parser)
            elif 'Unit.aspx' in conditional_part_url:
                # Если это страница с выбором юнита, выбираем и дальше парсим таблицу
                sub_details_js_list = parse_sub_details(conditional_part_url, parser)

                for sub_detail_js in sub_details_js_list:
                    parser.logger(f"[+]{cleared_model_name}::{issue_date}::{str(model_name)}", saveonly=False, first=False,
                           infunction=False)
                    # Переход на страницу поддеталей и парсинг таблицы
                    part_url = parser.selenium_crossing(conditional_part_url, sub_detail_js, driver)
                    if part_url is None:
                        continue

                    part_mage_url = parse_part_picture(URL, part_url, parser)
                    year_data.append(part_url)
                    year_data.append(part_mage_url)
                    cleared_model_name = clear_model_name_for_book(model_name, parser)
                    pass


def clear_model_name_for_book(model_name, parser):
    """
    Очищает имя модели для создания книги в Excel.
    Возвращает очищенное имя, которое содержит только английские буквы и цифры и имеет длину не более 30 символов.
    """
    try:
        # Удаление недопустимых символов, оставляя только английские буквы и цифры
        cleaned_name = re.sub(r'[^a-zA-Z0-9]', ' ', model_name)
        # Обрезка до 30 символов
        cleaned_name = cleaned_name[:30]
        # Удаление пробелов в начале и в конце
        cleaned_name = cleaned_name.strip()
        return cleaned_name
    except Exception as e:
        parser.logger(f"Ошибка обработки в функции clear_model_name_for_book: {e}")


def parse_complectation(url, driver, parser):
    """
    Нажимает на кнопку далее на старнице комплектаци
    """
    btn_selector = "#btnHyundaiNext"    #ID кнопки для перехода
    try:
        complectations_url = parser.selenium_click_and_get_page(url, btn_selector, driver)
        return complectations_url
    except Exception as e:
        parser.logger(f"Ошибка обработки в функици parse_complectation: {e}")

    
def parse_details_page(complectations_url, parser):
    """
    Парсит переход на страницу с таблицей с деталями
    Возвращает словарь деталей, где группа - ключ, а значение все js для перехода на таблицу с деталью этой группы
    """
    try:
        soup = BeautifulSoup(parser.fetch_data(complectations_url).text, 'lxml')
        next_sibling = soup.find('span', id="ctl00_cphMasterPage_trvGroups1").find_next_sibling()
        details_dict={}
        while next_sibling:
            if next_sibling.name == 'table':
                group_name = (next_sibling.find_all('a')[1]).text
                next_sibling = next_sibling.find_next_sibling()
                if next_sibling and next_sibling.name == 'div':
                    # Получение всех href, начинающихся с javascript
                    links = next_sibling.find_all('a', href=re.compile(r'^javascript'))
                    model_hrefs = [link.get('href').strip(';') for link in links]  # Список ссылок
                    # ЗАПОЛНЕНИЕ DICT
                    details_dict[group_name] = model_hrefs
            next_sibling = next_sibling.find_next_sibling()
        parser.save_data(name='Details', path='data', src=details_dict)
        return details_dict
    except Exception as e:
        parser.logger(f"Произошла ошибка в функции parse_details_page: {e}")


def parse_sub_details(part_url, parser):
    """
    Делает переход по юниту в селениуме
    Возвращает страницу part_url
    """

    """
    Возвращает js список для перехода на страницу таблицы подзапчасти
    """
    try:
        soup = BeautifulSoup(parser.fetch_data(part_url).text,'lxml')
        table = soup.find('table', id='ctl00_cphMasterPage_tblUnit')
        js_fns_a = table.find_all('a')  # Получение всех ссылок с js
        links_js = [js.get('href').strip(';') for js in js_fns_a]
        return links_js
    except Exception as e:
        parser.logger(f'Ошибка при выполнении функции parse_sub_detail: {e}', saveonly=False, first=False, infunction=True)



def parse_part_picture(gen_url, part_url, parser):
    """
    Парсит картинку part_image_url
    """

    try:
        soup = BeautifulSoup(parser.fetch_data(part_url).text, 'lxml')
        part_image_url = gen_url + str(soup.find('img', id='ctl00_cphMasterPage_imgParts')['src'])
        return part_image_url
    except Exception as e:
        parser.logger(f"Произошла ошибка в функции parse_part_picture: {e}")


def collect_playload(part_url, parser):
    """
    Функция открывает каждую группу деталей и парсит данные.
    Возвращает словарь с подставленными __VIEWSTATE, __VIEWSTATEGENERATOR, __EVENTVALIDATION
    Возрващает список id элементов, по которым нужно сделать запрос, чтобы открыть
    """
    try:
        form_data = {
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        '__VIEWSTATE': '',
        '__VIEWSTATEGENERATOR': '',
        'ctl00$cphMasterPage$urMain$UserOffice': 'ru',
        '__CALLBACKID': '__Page',
        '__CALLBACKPARAM': '',
        '__EVENTVALIDATION': ''
        }
        # Получаем __VIEWSTATEGENERATOR, __VIEWSTATE, __EVENTVALIDATION
        soup = BeautifulSoup(parser.fetch_data(part_url).text, 'lxml')
        viewstate = soup.find('input', {'name': '__VIEWSTATE'})['value']
        viewstate_generation = soup.find('input', {'name': '__VIEWSTATEGENERATOR'})['value']
        event_validation = soup.find('input', {'name': '__EVENTVALIDATION'})['value']

        form_data['__VIEWSTATE'] = viewstate
        form_data['__VIEWSTATEGENERATOR'] = viewstate_generation
        form_data['__EVENTVALIDATION'] = event_validation

        divs = soup.find_all(class_='CNode')
        ids = [div['id'] for div in divs]

        return form_data, ids
    except Exception as e:
        parser.logger(f"Произошла ошибка в функции collect_playload: {e}")


def parse_table(part_url, year_data, cleared_model_name, lock, parser):
    """
    Парсит таблицу деталей
    Возвращает список деталей (5шт)
    """
    try:
        # Собирает словарь playloads, парсит id товаров
        form_data, ids = collect_playload(part_url, parser)

        # Получаем
        response, session = parser.fetch_data(url=part_url, return_session=True)

        for i in range(len(ids)):
            form_data['__CALLBACKPARAM'] = ids[i]
            updated_response = parser.fetch_data(part_url, form_data, session)
            soup = BeautifulSoup(updated_response.text, 'lxml')
            rows = soup.find(class_='OpelParts').find_all('tr')[1:] # Пропускаем заголовок

            row_lst = []  # Список деталей для текущего ID
            for row in rows:
                # один ряд деталей
                details_list = [detail.text.strip() for detail in row]
                sorted_lst = [details_list[i] for i in [1, 4, 0, 2]]
                row_lst.append(sorted_lst)
            for row in row_lst:
                row_to_write = []
                row_to_write.extend(year_data)
                row_to_write.extend(row)
                parser.logger(f"[+]{row_to_write}", saveonly=False, first=False, infunction=True)
                add_to_new_sheet(str(parser.file_path), cleared_model_name, row_to_write, lock, parser)

    except Exception as e:
        parser.logger(f"Произошла ошибка в функции parse_table: {e}")


def create_new_book(file_path, sheet_name, parser):
    """
    Создает новую книгу и добавляет лист с указанным названием.
    """
    try:
        # Проверка, существует ли уже файл
        if os.path.exists(file_path):
            parser.logger(f"Файл {file_path} уже существует. Перезаписываем файл.")
        
        # Создание новой книги
        workbook = Workbook()
        sheet = workbook.active
        
        # Переименование листа
        sheet.title = sheet_name
        
        # Сохранение книги в файл
        workbook.save(file_path)
        parser.logger(f"Книга '{file_path}' с листом '{sheet_name}' успешно создана.")
    
    except Exception as e:
        parser.logger(f"Произошла ошибка при создании книги create_new_book: {e}")


def add_to_new_sheet(file_path, sheet_name, row, lock, parser):
    try:
        # Проверяем существование файла
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл по пути {file_path} не существует.")

        with lock:
            workbook = load_workbook(file_path)

        # Проверяем наличие листа
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                # Создаём новый лист, если не найден
                sheet = workbook.create_sheet(title=sheet_name)
                parser.logger(f'Лист с названием "{sheet_name}" был создан в книге по пути "{file_path}"')
            sheet.append(row)
            # Сохраняем изменения
            workbook.save(file_path)
            # parser.logger(f'Данные успешно добавлены в лист "{sheet_name}" в книге по пути "{file_path}"')

    except FileNotFoundError as fnf_error:
        parser.logger(f'Ошибка: {fnf_error}')
    except ValueError as val_error:
        parser.logger(f'Ошибка: {val_error}')
    except SheetTitleException as st_error:
        parser.logger(f'Ошибка: Недопустимое имя листа "{sheet_name}": {st_error}')
    except ValueError as val_error:
        parser.logger(f'Ошибка: Неверное значение в названии "{sheet_name}": {val_error}')
    except Exception as e:
        parser.logger(f'Ошибка добавления данных в лист "{sheet_name}" в функции add_to_new_sheet: {e}')

        
def parse():
    try:
        parser = Parser()
        parser.logger('|---Программа начала свою работу---|', False, True)
        driver = parser.setup_driver()
        time = Time()
        lock = parser.lock_treads
        # ---------------РАЗБОР МОДЕЛЕЙ НА СТРАНИЦЕ---------------
        parser.logger('|---Получение всех моделей...')
        model_name_for_print, MODEL_URL = choose_model_by_user(parse_all_models_into_file(MODELS_URL, parser))
        models_name, region_urls, years = parse_region(MODEL_URL, parser)
        parser.logger(f'|---Парсим детали {model_name_for_print}...', saveonly=False, first=False)

        # Добавляем книгу по названию модели
        file_path = os.path.join('data', 'BMW', f'{model_name_for_print}.xlsx')
        parser.file_path = file_path
        create_new_book(file_path=file_path, sheet_name=model_name_for_print, parser=parser)

        # Создание задания для поточной обработки
        tasks = [(model_name, region_url, year) for model_name, region_url, year in zip(models_name, region_urls, years)]
        THREADS = 12
        lock = Lock()
        with ThreadPoolExecutor(max_workers=THREADS) as executor:
            futures = [executor.submit(detail_process, task, lock, parser) for task in tasks]
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Ошибка потока: {e}")  
                
        time.end()
        driver.quit()
        parser.logger(f'|------------------------------------------------------|', saveonly=False, first=False, infunction=False)
        parser.logger(f'|---Модели были успешно собраны по пути data/BMW/output.xlsx.', saveonly=False, first=False, infunction=False)

    except KeyboardInterrupt:
        parser.logger('\nKeyboardInterrupt')
    except Exception as e:
        parser.logger(f'|---Ошибка в работе программы\n')
    finally:
        driver.quit()
        parser.logger(f'|------------------------------------------------------|', saveonly=False, first=False, infunction=False)
        parser.logger('|---Завершение работы программы...', saveonly=False, first=False, infunction=False)
    

if __name__ == "__main__":
    multiprocessing.set_start_method('spawn')
    parse()